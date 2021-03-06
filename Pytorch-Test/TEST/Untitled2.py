#!/usr/bin/env python
# coding: utf-8

# In[1]:


from torchvision import models
import torch
from PIL import Image
from torchvision import transforms
import pretrainedmodels
import time
import openpyxl as pyxl
import os, glob
import json
import argparse
from openpyxl.styles import Font


# In[2]:


use_GPU = True
model_label_path = "/media/Share/jingyun/Pytorch-training/files/snapshot/0920NEWFoodRecongnition-res50-classes.txt"


# In[3]:


device = torch.device("cuda:0" if torch.cuda.is_available() and use_GPU else "cpu")
label_dict = []
with open(model_label_path, "rt") as label_file:
    label_dict = [line.strip() for line in label_file.readlines()]
num_class = len(label_dict)
print("num_class:", num_class)


# In[4]:


model_name = "resnet50"
model_path = "/media/Share/jingyun/Pytorch-training/files/snapshot/BestWeight-0920NEWFoodRecongnition-res50-23-09-2019_15-38-00.pth"


# In[5]:


model_ft2 = pretrainedmodels.__dict__[model_name](num_classes=1000, pretrained='imagenet')
num_ftrs = model_ft2.last_linear.in_features
model_ft2.avgpool = torch.nn.AdaptiveAvgPool2d(1)
model_ft2.last_linear = torch.nn.Linear(num_ftrs, num_class)
print("loading in weights")
model_ft2.load_state_dict(torch.load(model_path,map_location='cpu'))
model_ft2.eval()
sm = torch.nn.Softmax(dim=1)


# In[6]:


val_transform = transforms.Compose([
        transforms.Resize(299),
        transforms.CenterCrop(224),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
    ])
model_ft2 = model_ft2.to(device)


# In[7]:


def dataset_list_JJ(path_to_dataset, keep_folder_id = []):
    ret_dict = {}
    for outer_folder in os.scandir(path_to_dataset):
        try:
            folder_id = outer_folder.name
        except Exception as e:
            print(outer_folder.name, e)
            folder_id = outer_folder.name
        print("processing",outer_folder.path,"folder_id:",folder_id)
        try:
            folder_id = int(folder_id)
        except:
            print(outer_folder.name,"cannot be change to an int")
        if len(keep_folder_id) and not str(folder_id) in keep_folder_id:
            print("skipping",outer_folder.path,"it is marked as to skip")
            continue
        print("including",outer_folder.name, outer_folder.path)
        ret_dict[outer_folder.name] = []
        for file in glob.iglob(outer_folder.path+'/**/*.jpg', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.jpeg', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.JPEG', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.JPG', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.png', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
    return ret_dict

def pil_loader(path):
    try:
        with open(path, 'rb') as f:
            img = Image.open(f)
            return img.convert('RGB')
    except Exception as e:
        print(e, "img_path:", path)

val_dict_img_list = dataset_list_JJ("/media/Share/jingyun/data/Totest_NBF_general")


# In[8]:


total_pic=0
total_top1=0
total_top5=0
total_top25=0

food_results = {}
for food_label, img_list in val_dict_img_list.items():
    time_since = time.time()
    try:
        food_label = food_label.strip() #去除首尾空格
        food_label_idx = label_dict.index(food_label) #从列表中找出某个值第一个匹配项的索引位置
    except Exception as e:
        print("Not found! Skipping this food entirely",e)
        food_label_idx = len(label_dict)
        continue
    print("Validating food:", food_label)
    food_results[food_label] = {"top1":0,"top5":0,"top25":0,"foundIn":[],"recogResult":{},"img_tested":[],"total":len(img_list),"foodaccuracy_top1":0,                "foodaccuracy_top5":0,"foodaccuracy_top25":0}
    for img_path in img_list:
        if "@eaDir" in img_path:
            continue
        input_tensor = val_transform(pil_loader(img_path)).unsqueeze(0) #载入图片并进行预处理，对数据维度进行扩充。给指定位置加上维数为一的维度
        input_t = torch.autograd.Variable(input_tensor, requires_grad = False) #封装成Variable
        input_t = input_t.to(device)
        output_logits = model_ft2(input_t) #得到输出结果
        prob = sm(output_logits) #对输出进行softmax操作
        prob, cls_id = prob.topk(k=len(label_dict), dim=1)
        cls_list = cls_id.squeeze(0).tolist()
        prob_list = prob.squeeze(0).tolist()
        foundIn =  cls_list.index(food_label_idx)
        food_results[food_label]["foundIn"].append(foundIn)
        food_results[food_label]["img_tested"].append(img_path)
        if foundIn <= 24:
            food_results[food_label]["top25"] += 1
        if foundIn <= 4:
            food_results[food_label]["top5"] += 1
        if foundIn == 0:
            food_results[food_label]["top1"] += 1
        if not label_dict[cls_list[0]] in food_results[food_label]["recogResult"]:
            food_results[food_label]["recogResult"][label_dict[cls_list[0]]] = 0
        food_results[food_label]["recogResult"][label_dict[cls_list[0]]] += 1
    print("Complete food: {} in:".format(food_label), time.time() - time_since)
    if food_results[food_label]["total"]:
        food_results[food_label]["foodaccuracy_top1"]=food_results[food_label]["top1"]/food_results[food_label]["total"]
        food_results[food_label]["foodaccuracy_top5"]=food_results[food_label]["top5"]/food_results[food_label]["total"]
        food_results[food_label]["foodaccuracy_top25"]=food_results[food_label]["top25"]/food_results[food_label]["total"]   
    else:
        food_results[food_label]["foodaccuracy_top1"]=0
        food_results[food_label]["foodaccuracy_top5"]=0
        food_results[food_label]["foodaccuracy_top25"]=0
    total_pic+=food_results[food_label]["total"]
    total_top1+=food_results[food_label]["top1"]
    total_top5+=food_results[food_label]["top5"]
    total_top25+=food_results[food_label]["top25"]
print("Done!")


# In[9]:


import openpyxl as pyxl
from openpyxl.drawing.image import Image as pyxlImg
from openpyxl.utils import get_column_letter
import os

errAnaWorkbook = pyxl.Workbook()
ws = errAnaWorkbook.active
ws.title = 'Err Report-bothTrials'
ws.cell(row=1,column=1,value="food_label")
ws.cell(row=1,column=2,value="foodaccuracy_top1")
ws.cell(row=1,column=3,value="foodaccuracy_top5")
ws.cell(row=1,column=4,value="foodaccuracy_top25")
ws.cell(row=1,column=5,value="Pic_NUM")
resStartRow = 1
for p in range(0,num_class):
    resStartRow += 1
    ws.cell(row=resStartRow,column=1,value=label_dict[p])
    ws.cell(row=resStartRow,column=2,value=str(food_results[label_dict[p]]["foodaccuracy_top1"]))
    ws.cell(row=resStartRow,column=3,value=str(food_results[label_dict[p]]["foodaccuracy_top5"]))
    ws.cell(row=resStartRow,column=4,value=str(food_results[label_dict[p]]["foodaccuracy_top25"]))
    ws.cell(row=resStartRow,column=5,value=str(food_results[label_dict[p]]["total"]))
ws.cell(row=resStartRow+1,column=1,value="Total accuracy")
ws.cell(row=resStartRow+1,column=2,value=str(total_top1/total_pic))
ws.cell(row=resStartRow+1,column=3,value=str(total_top5/total_pic))
ws.cell(row=resStartRow+1,column=4,value=str(total_top25/total_pic))
errAnaWorkbook.save("G_0925.xlsx")    


# In[10]:


#print(str(food_results['12_Rice']["foodaccuracy_top25"]))
#print(str(food_results['12_Rice']["foodaccuracy_top5"]))
#print(str(food_results['59_Banana']["foodaccuracy_top1"]))
#print(str(total_pic))
#print(str(food_results['12_Rice']["top1"]))
#print(str(food_results['12_Rice']["top5"]))
#print(str(food_results['12_Rice']["top25"]))
#print(str(food_results[food_label]["total"]))


# In[ ]:




