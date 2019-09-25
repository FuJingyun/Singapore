#!/usr/bin/env python
# coding: utf-8

# In[1]:


from torchvision import models
import torch
from PIL import Image
from torchvision import transforms
import pretrainedmodels
import time
import openpyxl as pyxl
import os, glob
import json
import argparse
from openpyxl.styles import Font


# In[2]:


use_GPU = True
model_label_path = "/media/Share/jingyun/Pytorch-training/files/snapshot/foodNonfood-res50-classes.txt"


# In[3]:


device = torch.device("cuda:0" if torch.cuda.is_available() and use_GPU else "cpu")
label_dict = []
with open(model_label_path, "rt") as label_file:
    label_dict = [line.strip() for line in label_file.readlines()]
num_class = len(label_dict)
print("num_class:", num_class)


# In[4]:


model_name = "resnet50"
model_path = "/media/Share/jingyun/Pytorch-training/files/snapshot/BestWeight-foodNonfood-res50-19-08-2019_13-23-22.pth"


# In[5]:


model_ft2 = pretrainedmodels.__dict__[model_name](num_classes=1000, pretrained='imagenet')
num_ftrs = model_ft2.last_linear.in_features
model_ft2.avgpool = torch.nn.AdaptiveAvgPool2d(1)
model_ft2.last_linear = torch.nn.Linear(num_ftrs, num_class)
print("loading in weights")
model_ft2.load_state_dict(torch.load(model_path,map_location='cpu'))
model_ft2.eval()
sm = torch.nn.Softmax(dim=1)


# In[6]:


val_transform = transforms.Compose([
        transforms.Resize(299),
        transforms.CenterCrop(224),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
    ])
model_ft2 = model_ft2.to(device)


# In[7]:


def dataset_list_JJ(path_to_dataset, keep_folder_id = []):
    ret_dict = {}
    for outer_folder in os.scandir(path_to_dataset):
        try:
            folder_id = outer_folder.name
        except Exception as e:
            print(outer_folder.name, e)
            folder_id = outer_folder.name
        print("processing",outer_folder.path,"folder_id:",folder_id)
        try:
            folder_id = int(folder_id)
        except:
            print(outer_folder.name,"cannot be change to an int")
        if len(keep_folder_id) and not str(folder_id) in keep_folder_id:
            print("skipping",outer_folder.path,"it is marked as to skip")
            continue
        print("including",outer_folder.name, outer_folder.path)
        ret_dict[outer_folder.name] = []
        for file in glob.iglob(outer_folder.path+'/**/*.jpg', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.jpeg', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.JPEG', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.JPG', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
        for file in glob.iglob(outer_folder.path+'/**/*.png', recursive=True):
            if not "@eaDir" in file:
                ret_dict[outer_folder.name].append(file)
    return ret_dict

def pil_loader(path):
    try:
        with open(path, 'rb') as f:
            img = Image.open(f)
            return img.convert('RGB')
    except Exception as e:
        print(e, "img_path:", path)

val_dict_img_list = dataset_list_JJ("/media/Share/jingyun/data/test_living")


# In[8]:


food_results = {}
for food_label, img_list in val_dict_img_list.items():
    time_since = time.time()
    try:
        food_label = food_label.strip()
        food_label_idx = label_dict.index(food_label)
    except Exception as e:
        print("Not found! Skipping this food entirely",e)
        food_label_idx = len(label_dict)
        continue
    print("Validating food:", food_label)
    food_results[food_label] = {"top1":0,"top5":0,"top25":0,"foundIn":[],"recogResult":{},"img_tested":[],"total":len(img_list)}
    for img_path in img_list:
        if "@eaDir" in img_path:
            continue
        input_tensor = val_transform(pil_loader(img_path)).unsqueeze(0)
        input_t = torch.autograd.Variable(input_tensor, requires_grad = False)
        input_t = input_t.to(device)
        output_logits = model_ft2(input_t)
        prob = sm(output_logits)
        prob, cls_id = prob.topk(k=len(label_dict), dim=1)
        cls_list = cls_id.squeeze(0).tolist()
        prob_list = prob.squeeze(0).tolist()
        foundIn =  cls_list.index(food_label_idx)
        food_results[food_label]["foundIn"].append(foundIn)
        food_results[food_label]["img_tested"].append(img_path)
        if foundIn <= 24:
            food_results[food_label]["top25"] += 1
        if foundIn <= 4:
            food_results[food_label]["top5"] += 1
        if foundIn == 0:
            food_results[food_label]["top1"] += 1
        if not label_dict[cls_list[0]] in food_results[food_label]["recogResult"]:
            food_results[food_label]["recogResult"][label_dict[cls_list[0]]] = 0
        food_results[food_label]["recogResult"][label_dict[cls_list[0]]] += 1
    print("Complete food: {} in:".format(food_label), time.time() - time_since)
print("Done!")


# In[9]:


#food_results


# In[10]:


from PIL import Image

food_results = []
for food_label, img_list in val_dict_img_list.items():
    time_since = time.time()
    class_label_indices = []
    try:
        food_label = food_label.strip()
        food_label_idx = label_dict.index(food_label)
    except Exception as e:
        print("Not found! Skipping this food entirely",e)
        food_label_idx = len(label_dict)
        continue
    print("Validating food:", food_label)
    
    for img_path in img_list:
        if "@eaDir" in img_path:
            continue
        input_tensor = val_transform(pil_loader(img_path)).unsqueeze(0)
        input_t = torch.autograd.Variable(input_tensor, requires_grad = False)
        input_t = input_t.to(device)
        output_logits = model_ft2(input_t)
        prob = sm(output_logits)
        prob, cls_id = prob.topk(k=len(label_dict), dim=1)
        cls_list = cls_id.squeeze(0).tolist()
        prob_list = prob.squeeze(0).tolist()
        foundIn = num_class
        firstTen = []
        for i in range(len(cls_list)):
            firstTen.append({"name":label_dict[cls_list[i]], "confidence":prob_list[i]})
        foundIn = cls_list.index(food_label_idx)
        img_obj = Image.open(img_path)
        img_obj = img_obj.convert("RGB")
        img_obj.save(img_path)
        food_results.append({"ipath":img_path,"GTs": food_label, "foundIn": foundIn, "first 10 results":firstTen})
    print("Complete food: {} in:".format(food_label), time.time() - time_since)
print("Done!")


# In[11]:


import openpyxl as pyxl
from openpyxl.drawing.image import Image as pyxlImg
from openpyxl.utils import get_column_letter
import os

errAnaWorkbook = pyxl.Workbook()
ws = errAnaWorkbook.active
ws.title = 'Err Report-bothTrials'
ws.cell(row=1,column=1,value="Filename")
ws.cell(row=1,column=2,value="Image")
ws.cell(row=1,column=3,value="Ground truth(s)")
ws.cell(row=1,column=4,value="Found in")
ws.cell(row=1,column=5,value="First 10 results")
resStartRow = 1
ws.column_dimensions[chr(ord('A')+1)].width = 18

for entry in food_results:
    resStartCol = 5
    resStartRow += 1
    testImg = pyxlImg(entry["ipath"])
    testImg.width = 120
    testImg.height = 120
    testImg.anchor = '{}{}'.format(get_column_letter(2),resStartRow)
    ws.add_image(testImg)
    ws.row_dimensions[resStartRow].height = 95
    ws.cell(row=resStartRow,column=1,value="{}".format(os.path.basename(entry["ipath"])))
    ws.cell(row=resStartRow,column=3,value="{}".format(entry["GTs"]))
    ws.cell(row=resStartRow,column=4,value="=value({})".format(entry["foundIn"]+1))
    for res in entry["first 10 results"]:
        ws.cell(row=resStartRow,column=resStartCol,value="{} ({:0.3f})".format(res["name"], res["confidence"]))
        resStartCol += 1
resStartRow += 1
#ws.cell(row=resStartRow,column=3,value="Top-1")
#ws.cell(row=resStartRow,column=4,value="=COUNTIF($D$2:$D${0},1)/COUNTA($D$2:$D${0})".format(resStartRow-1))
resStartRow += 1
errAnaWorkbook.save("test_living0820.xlsx")


# In[ ]:




